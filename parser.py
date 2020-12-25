import pandas as pd
from collections import Counter
import openpyxl
import ast
wb = openpyxl.Workbook()
wb.create_sheet(title = 'List1')
sheet = wb['List1']
df = pd.read_csv("PlayApps_202010.csv")
saved_column = df["Amount (Merchant Currency)"]
summ = []
comiss = []
pairs = []
 #you can also use df['column_name']
k = 0
flag = 0
for i in saved_column:    
    if k % 2 == 0:
        if i > 0:
            if flag == 1:
                comiss.append(0)
            summ.append(float(i))
            flag = 1
        else:
            comiss.append(float(i))
            flag = 0    
        
    else:
        if i < 0:
            comiss.append(float(i))
            flag = 0
        else:
            if flag == 1:
                comiss.append(0)
            summ.append(float(i))
            flag = 1
    
    if k==len(saved_column)-1 and i>0:
        comiss.append(0)



    k += 1

print(k)
print(summ)
print(comiss)


for i in range(len(summ)):
    pairs.append([summ[i], comiss[i]])



sheet.cell(row = 1, column = 1).value = "A"
sheet.cell(row = 1, column = 2).value = "Sum from table"
sheet.cell(row = 1, column = 3).value = "Comission from table"
sheet.cell(row = 1, column = 4).value = "Qty"
sheet.cell(row = 1, column = 5).value = "Unit Price"
sheet.cell(row = 1, column = 6).value = "Amount"
a = Counter(str(e) for e in pairs)
ind = 2
summa = 0
for i in a:
    # sheet.cell(row = ind, column = 1) = ind
    # sheet.cell(row = ind, column = 1) = ind
    ls = ast.literal_eval(i)
    print(ls[0], ls[1], a[i], round(ls[0]+ls[1],2), round(a[i]*(ls[0]+ls[1]),2))
    sheet.cell(row = ind, column = 1).value = ind
    sheet.cell(row = ind, column = 2).value = ls[0]
    sheet.cell(row = ind, column = 3).value = ls[1]
    sheet.cell(row = ind, column = 4).value = a[i]
    sheet.cell(row = ind, column = 5).value = "$ "+str(round(ls[0]+ls[1],2))
    sheet.cell(row = ind, column = 6).value = "$ "+str(round(a[i]*(ls[0]+ls[1]),2))
    ind += 1
    summa += round(a[i]*(ls[0]+ls[1]),2)
sheet.cell(row = ind, column = 1).value = ind    
sheet.cell(row = ind, column = 6).value = "$ "+str(summa)
wb.save('resultat.xlsx')





    
